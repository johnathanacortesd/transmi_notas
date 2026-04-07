import streamlit as st
import pandas as pd
from openpyxl import load_workbook
import datetime
import joblib
import numpy as np
import dossier_utils as utils

# --- Configuración de la página ---
st.set_page_config(
    page_title="Dossier Intelligence · Transmilenio",
    page_icon="🚌",
    layout="wide",
    initial_sidebar_state="collapsed"
)

st.markdown("""
<style>
    @import url('https://fonts.googleapis.com/css2?family=DM+Sans:wght@300;400;500;600;700&family=DM+Mono:wght@400;500&display=swap');

    html, body, [class*="css"] { font-family: 'DM Sans', sans-serif; }
    .stApp { background-color: #F7F8FA; }
    #MainMenu, footer, header { visibility: hidden; }
    .block-container {
        padding-top: 2rem;
        padding-bottom: 3rem;
        max-width: 1100px;
    }

    .app-header {
        background: linear-gradient(135deg, #B71C1C 0%, #D32F2F 60%, #F44336 100%);
        border-radius: 12px;
        padding: 1.1rem 1.8rem;
        margin-bottom: 1.2rem;
        position: relative;
        overflow: hidden;
        display: flex;
        align-items: center;
        gap: 1.2rem;
    }
    .app-header::before {
        content: '';
        position: absolute;
        top: -30px; right: -30px;
        width: 120px; height: 120px;
        background: rgba(255,255,255,0.06);
        border-radius: 50%;
    }
    .app-header h1 { display: none; }
    .app-header p {
        color: rgba(255,255,255,0.9);
        font-size: 0.92rem;
        font-weight: 500;
        margin: 0;
    }
    .app-header .badge {
        display: inline-block;
        background: rgba(255,255,255,0.15);
        color: #fff;
        font-size: 0.7rem;
        font-weight: 600;
        letter-spacing: 0.08em;
        text-transform: uppercase;
        padding: 0.2rem 0.6rem;
        border-radius: 20px;
        border: 1px solid rgba(255,255,255,0.25);
        white-space: nowrap;
        flex-shrink: 0;
    }

    .card {
        background: #FFFFFF;
        border: 1px solid #E8ECEF;
        border-radius: 12px;
        padding: 1.6rem 1.8rem;
        margin-bottom: 1.2rem;
        box-shadow: 0 1px 4px rgba(183,28,28,0.06);
    }
    .card-title {
        font-size: 0.8rem;
        font-weight: 600;
        letter-spacing: 0.1em;
        text-transform: uppercase;
        color: #D32F2F;
        margin-bottom: 1rem;
    }

    .metrics-row {
        display: flex;
        gap: 1rem;
        margin: 1.5rem 0;
    }
    .metric-card {
        flex: 1;
        background: #FFFFFF;
        border: 1px solid #E8ECEF;
        border-radius: 12px;
        padding: 1.4rem 1.6rem;
        text-align: center;
        box-shadow: 0 1px 4px rgba(183,28,28,0.06);
    }
    .metric-card .metric-value {
        font-size: 2.2rem;
        font-weight: 700;
        color: #B71C1C;
        line-height: 1;
        margin-bottom: 0.4rem;
        font-family: 'DM Mono', monospace;
    }
    .metric-card .metric-label {
        font-size: 0.78rem;
        color: #8F6B6B;
        font-weight: 500;
        letter-spacing: 0.03em;
    }
    .metric-card.accent .metric-value { color: #D32F2F; }
    .metric-card.muted  .metric-value { color: #AC9B9B; }

    .success-banner {
        background: linear-gradient(135deg, #B71C1C, #D32F2F);
        border-radius: 12px;
        padding: 1.2rem 1.8rem;
        display: flex;
        align-items: center;
        gap: 1rem;
        margin: 1.5rem 0;
        box-shadow: 0 2px 12px rgba(183,28,28,0.2);
    }
    .success-banner .icon { font-size: 1.6rem; line-height: 1; }
    .success-banner .text strong {
        display: block;
        color: #FFFFFF;
        font-size: 1rem;
        font-weight: 700;
        margin-bottom: 0.15rem;
    }
    .success-banner .text span {
        color: rgba(255,255,255,0.72);
        font-size: 0.85rem;
    }

    .step {
        display: flex;
        align-items: flex-start;
        gap: 1rem;
        margin-bottom: 0.8rem;
    }
    .step-num {
        min-width: 28px; height: 28px;
        background: #D32F2F;
        color: white;
        border-radius: 50%;
        display: flex;
        align-items: center;
        justify-content: center;
        font-size: 0.78rem;
        font-weight: 700;
        margin-top: 1px;
    }
    .step-text { color: #5A2D2D; font-size: 0.9rem; line-height: 1.6; }

    [data-testid="stFileUploader"] {
        background: #FFFFFF;
        border: 2px dashed #D4B2B2;
        border-radius: 12px;
        padding: 0.5rem;
        transition: border-color 0.2s;
    }
    [data-testid="stFileUploader"]:hover { border-color: #D32F2F; }

    .stButton > button[kind="primary"] {
        background: linear-gradient(135deg, #B71C1C, #D32F2F);
        color: white;
        border: none;
        border-radius: 10px;
        padding: 0.7rem 2rem;
        font-family: 'DM Sans', sans-serif;
        font-size: 0.95rem;
        font-weight: 600;
        letter-spacing: 0.01em;
        transition: all 0.2s;
        box-shadow: 0 2px 8px rgba(183,28,28,0.25);
        width: 100%;
        height: 48px;
    }
    .stButton > button[kind="primary"]:hover {
        background: linear-gradient(135deg, #8E1515, #B71C1C);
        box-shadow: 0 4px 14px rgba(183,28,28,0.35);
        transform: translateY(-1px);
    }
    .stButton > button[kind="primary"]:disabled {
        background: #D4C5C5;
        box-shadow: none;
        transform: none;
    }

    .stDownloadButton > button {
        background: #FFFFFF;
        color: #B71C1C;
        border: 2px solid #D32F2F;
        border-radius: 10px;
        padding: 0.65rem 1.8rem;
        font-family: 'DM Sans', sans-serif;
        font-size: 0.9rem;
        font-weight: 600;
        transition: all 0.2s;
        width: 100%;
        height: 48px;
    }
    .stDownloadButton > button:hover {
        background: #FEF2F2;
        box-shadow: 0 2px 10px rgba(183,28,28,0.15);
        transform: translateY(-1px);
    }

    .stWarning { background: #FFF8ED; border: 1px solid #F5C97A; border-radius: 10px; }
    .stError   { background: #FEF2F2; border: 1px solid #FCA5A5; border-radius: 10px; }

    .stProgress > div > div > div {
        background: linear-gradient(90deg, #D32F2F, #F44336);
        border-radius: 4px;
    }

    .streamlit-expanderHeader {
        background: #F7F8FA;
        border-radius: 8px;
        font-weight: 500;
        color: #B71C1C;
    }

    [data-testid="stDataFrame"] {
        border-radius: 10px;
        overflow: hidden;
        border: 1px solid #E8ECEF;
    }

    hr { border: none; border-top: 1px solid #E8ECEF; margin: 1.5rem 0; }
    .stSpinner > div { border-top-color: #D32F2F !important; }

    .file-status {
        display: flex;
        align-items: center;
        gap: 0.6rem;
        padding: 0.7rem 1rem;
        border-radius: 8px;
        font-size: 0.87rem;
        font-weight: 500;
        margin-top: 0.5rem;
    }
    .file-status.ok      { background: #FEF2F2; color: #B71C1C; border: 1px solid #D4B2B2; }
    .file-status.missing { background: #FFF8ED; color: #92400E; border: 1px solid #F5C97A; }

    .results-header {
        font-size: 1.1rem;
        font-weight: 700;
        color: #B71C1C;
        margin: 1.8rem 0 1rem 0;
        padding-bottom: 0.5rem;
        border-bottom: 2px solid #E8ECEF;
    }

    .stTabs [data-baseweb="tab-list"] { gap: 0.5rem; }
    .stTabs [data-baseweb="tab"] {
        font-family: 'DM Sans', sans-serif;
        font-weight: 600;
        font-size: 0.9rem;
        color: #8F6B6B;
        border-radius: 8px 8px 0 0;
        padding: 0.6rem 1.4rem;
    }
    .stTabs [aria-selected="true"] {
        color: #B71C1C !important;
        border-bottom: 3px solid #D32F2F;
    }
</style>
""", unsafe_allow_html=True)


# ==============================================================================
# CONSTANTES COMPARTIDAS
# ==============================================================================

FINAL_ORDER = [
    "ID Noticia", "Fecha", "Hora", "Medio", "Tipo de Medio", "Sección - Programa",
    "Región", "Título", "Autor - Conductor", "Nro. Pagina", "Dimensión",
    "Duración - Nro. Caracteres", "CPE", "Tier", "Audiencia", "Tono",
    "Temas Generales - Tema", "Resumen - Aclaracion", "Link Nota",
    "Link (Streaming - Imagen)", "Menciones - Empresa"
]

TIPO_MEDIO_MAP = {
    'online': 'Internet', 'diario': 'Prensa',
    'am': 'Radio', 'fm': 'Radio',
    'aire': 'Televisión', 'cable': 'Televisión',
    'revista': 'Revistas'
}


# ==============================================================================
# FUNCIONES DE CARGA Y TRANSFORMACIÓN
# ==============================================================================

@st.cache_resource
def load_ml_models():
    try:
        sentiment_pipeline = joblib.load('pipeline_sentimiento_tm.pkl')
        topic_pipeline = joblib.load('pipeline_tema_tm.pkl')
        return sentiment_pipeline, topic_pipeline
    except FileNotFoundError as e:
        st.error(
            f"**Error Crítico:** No se encontró `{e.filename}`. "
            "Asegúrate de que los archivos .pkl estén en la misma carpeta que app.py."
        )
        st.stop()


def read_dossier(dossier_file):
    wb = load_workbook(dossier_file)
    sheet = wb.active
    original_headers = [cell.value for cell in sheet[1] if cell.value]

    rows_data = []
    for row in sheet.iter_rows(min_row=2, values_only=False):
        if all(c.value is None for c in row):
            continue
        row_values = {}
        for i, header in enumerate(original_headers):
            if header in ['Link Nota', 'Link (Streaming - Imagen)']:
                row_values[header] = utils.extract_link_from_cell(row[i])
            else:
                row_values[header] = row[i].value
        rows_data.append(row_values)

    return pd.DataFrame(rows_data)


def load_config_maps(config_file, include_mentions=False):
    try:
        config_sheets = pd.read_excel(config_file, sheet_name=None)

        region_map = pd.Series(
            config_sheets['Regiones'].iloc[:, 1].values,
            index=config_sheets['Regiones'].iloc[:, 0].astype(str).str.lower().str.strip()
        ).to_dict()

        internet_map = pd.Series(
            config_sheets['Internet'].iloc[:, 1].values,
            index=config_sheets['Internet'].iloc[:, 0].astype(str).str.lower().str.strip()
        ).to_dict()

        if include_mentions:
            mention_map = pd.Series(
                config_sheets['Menciones'].iloc[:, 1].values,
                index=config_sheets['Menciones'].iloc[:, 0].astype(str).str.strip()
            ).to_dict()
            return region_map, internet_map, mention_map

        return region_map, internet_map

    except Exception as e:
        st.error(f"**Error al cargar `Configuracion.xlsx`:** {e}")
        st.stop()


def apply_common_transformations(df, region_map, internet_map):
    if 'Fecha' in df.columns:
        df['Fecha'] = pd.to_datetime(df['Fecha'], errors='coerce', dayfirst=True)

    if 'Título' in df.columns:
        df['Título'] = df['Título'].apply(utils.clean_title)

    if 'Resumen - Aclaracion' in df.columns:
        df['Resumen - Aclaracion'] = df['Resumen - Aclaracion'].apply(utils.corregir_resumen)

    df['Tipo de Medio'] = (
        df['Tipo de Medio'].str.lower().str.strip()
        .map(TIPO_MEDIO_MAP)
        .fillna(df['Tipo de Medio'])
    )

    if 'Medio' in df.columns:
        df['Región'] = df['Medio'].astype(str).str.lower().str.strip().map(region_map)

    is_internet = df['Tipo de Medio'] == 'Internet'
    if is_internet.any():
        df.loc[is_internet, 'Medio'] = (
            df.loc[is_internet, 'Medio'].astype(str).str.lower().str.strip()
            .map(internet_map)
            .fillna(df.loc[is_internet, 'Medio'])
        )

    is_print     = df['Tipo de Medio'].isin(['Prensa', 'Revistas'])
    is_broadcast = df['Tipo de Medio'].isin(['Radio', 'Televisión'])

    if 'Link Nota' in df.columns and 'Link (Streaming - Imagen)' in df.columns:
        df.loc[is_internet, ['Link Nota', 'Link (Streaming - Imagen)']] = (
            df.loc[is_internet, ['Link (Streaming - Imagen)', 'Link Nota']].values
        )
        cond_copy = is_print & df['Link Nota'].isnull() & df['Link (Streaming - Imagen)'].notnull()
        df.loc[cond_copy, 'Link Nota'] = df.loc[cond_copy, 'Link (Streaming - Imagen)']
        df.loc[is_print | is_broadcast, 'Link (Streaming - Imagen)'] = None

    if 'Duración - Nro. Caracteres' in df.columns and 'Dimensión' in df.columns:
        df.loc[is_broadcast, 'Dimensión'] = df.loc[is_broadcast, 'Duración - Nro. Caracteres']
        df.loc[is_broadcast, 'Duración - Nro. Caracteres'] = np.nan

    return df


def apply_mention_mapping(df, mention_map):
    if 'Menciones - Empresa' not in df.columns:
        return df

    def map_mentions_cell(cell_value):
        if not isinstance(cell_value, str) or not cell_value.strip():
            return cell_value
        parts = [p.strip() for p in cell_value.split(';')]
        mapped_parts = [mention_map.get(p, p) for p in parts if p]
        return '; '.join(mapped_parts)

    df['Menciones - Empresa'] = df['Menciones - Empresa'].apply(map_mentions_cell)
    return df


# ==============================================================================
# PESTAÑA 1: PROCESO COMPLETO 
# ==============================================================================

def run_full_process(dossier_file, config_file, download_placeholder):
    st.markdown("<hr>", unsafe_allow_html=True)
    progress_bar = st.progress(0, text="Iniciando proceso...")

    progress_bar.progress(5, text="Paso 1 / 7 — Cargando modelos y configuración...")
    sentiment_pipeline, topic_pipeline = load_ml_models()
    region_map, internet_map = load_config_maps(config_file, include_mentions=False)

    progress_bar.progress(15, text="Paso 2 / 7 — Leyendo Dossier...")
    df = read_dossier(dossier_file)

    progress_bar.progress(25, text="Paso 3 / 7 — Aplicando mapeos y normalizaciones...")
    df = apply_common_transformations(df, region_map, internet_map)

    progress_bar.progress(45, text="Paso 4 / 7 — Detectando duplicados...")
    df = utils.detect_duplicates_optimized(df)

    progress_bar.progress(65, text="Paso 5 / 7 — Aplicando modelos de IA a noticias únicas...")
    
    # Aislar SOLO las noticias únicas para procesarlas
    df_valid = df[~df['is_duplicate']].copy()
    
    if not df_valid.empty:
        # 1. Preparar texto
        df_valid['texto_crudo'] = df_valid['Título'].fillna('') + ' ' + df_valid['Resumen - Aclaracion'].fillna('')
        df_valid['texto_limpio_ia'] = df_valid['texto_crudo'].apply(utils.limpiar_texto)
        
        # 2. Predecir Sentimiento
        preds_sent = sentiment_pipeline.predict(df_valid['texto_limpio_ia'])
        def capitalizar_tono(p):
            if str(p).lstrip('-').isdigit(): 
                return {2: 'Positivo', 1: 'Neutro', 0: 'Negativo', -1: 'Negativo'}.get(int(p), 'Indefinido')
            return str(p).capitalize()
        df_valid['Tono'] = [capitalizar_tono(p) for p in preds_sent]
        
        # 3. Predecir Tema
        df_valid['Temas Generales - Tema'] = topic_pipeline.predict(df_valid['texto_limpio_ia'])

        # 4. HOMOGENEIZAR TEMA Y TONO EN NOTICIAS SIMILARES NO-DUPLICADAS
        progress_bar.progress(80, text="Paso 6 / 7 — Homogeneizando Tono y Tema en noticias similares...")
        
        def generar_clave_similitud(row):
            titulo = utils.normalize_title_for_comparison(row.get('Título', ''))
            if len(titulo) > 10: return titulo
            resumen = str(row.get('texto_limpio_ia', ''))[:80]
            return resumen if resumen else f"id_{row.name}"
            
        df_valid['clave_similitud'] = df_valid.apply(generar_clave_similitud, axis=1)
        
        # Asignar la moda al grupo
        df_valid['Tono'] = df_valid.groupby('clave_similitud')['Tono'].transform(lambda x: x.mode()[0] if not x.mode().empty else x)
        df_valid['Temas Generales - Tema'] = df_valid.groupby('clave_similitud')['Temas Generales - Tema'].transform(lambda x: x.mode()[0] if not x.mode().empty else x)

        # Actualizar dataframe original con los datos válidos homogeneizados
        df.update(df_valid[['Tono', 'Temas Generales - Tema']])

    # 5. NUEVA REGLA DE NEGOCIO: "TITULARES"
    progress_bar.progress(90, text="Aplicando regla de negocio: 'Titulares'...")
    if 'Título' in df.columns:
        mask_titulares = df['Título'].astype(str).str.contains('titulares', case=False, na=False)
        df.loc[mask_titulares, 'Tono'] = 'Neutro'
        df.loc[mask_titulares, 'Temas Generales - Tema'] = 'Entorno e información general'

    # 6. RESTAURAR EL COMPORTAMIENTO ORIGINAL PARA LAS DUPLICADAS
    # (Hacerlo al final garantiza que si hay un Titular duplicado, diga "Duplicada")
    mask_dup = df['is_duplicate']
    if mask_dup.any():
        if 'Temas Generales - Tema' in df.columns:
            df.loc[mask_dup, 'Temas Generales - Tema'] = '-'
        df.loc[mask_dup, 'Tono'] = 'Duplicada'

    progress_bar.progress(100, text="✓ Proceso completado")

    total        = len(df)
    dups_count   = int(mask_dup.sum())
    unique_count = total - dups_count
    filename     = f"Dossier_Transmilenio_{datetime.datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    excel_data   = utils.to_excel_from_df(df, FINAL_ORDER)

    with download_placeholder:
        st.download_button(
            label="⬇ Descargar archivo procesado (.xlsx)",
            data=excel_data,
            file_name=filename,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    st.markdown(f"""
    <div class="success-banner">
        <div class="icon">🚌</div>
        <div class="text">
            <strong>Proceso finalizado correctamente</strong>
            <span>{total:,} filas procesadas · {unique_count:,} únicas · {dups_count:,} duplicadas</span>
        </div>
    </div>
    """, unsafe_allow_html=True)

    st.markdown('<p class="results-header">Resumen del proceso</p>', unsafe_allow_html=True)
    st.markdown(f"""
    <div class="metrics-row">
        <div class="metric-card">
            <div class="metric-value">{total:,}</div>
            <div class="metric-label">Filas totales procesadas</div>
        </div>
        <div class="metric-card accent">
            <div class="metric-value">{unique_count:,}</div>
            <div class="metric-label">Noticias únicas analizadas</div>
        </div>
        <div class="metric-card muted">
            <div class="metric-value">{dups_count:,}</div>
            <div class="metric-label">Filas marcadas como duplicadas</div>
        </div>
    </div>
    """, unsafe_allow_html=True)

    st.markdown('<p class="results-header">Previsualización de resultados</p>', unsafe_allow_html=True)
    final_cols_in_df = [col for col in FINAL_ORDER if col in df.columns]
    df_display = df[final_cols_in_df].copy()
    if 'Fecha' in df_display.columns:
        df_display['Fecha'] = (
            pd.to_datetime(df_display['Fecha'])
            .dt.strftime('%d/%m/%Y')
            .replace('NaT', 'FECHA INVÁLIDA')
        )
    st.dataframe(df_display, use_container_width=True, hide_index=True)


# ==============================================================================
# PESTAÑA 2: EXPANDIR POR MENCIONES (sin IA, con mapeo de menciones)
# ==============================================================================

def run_expand_process(dossier_file, config_file, download_placeholder):
    st.markdown("<hr>", unsafe_allow_html=True)
    progress_bar = st.progress(0, text="Iniciando proceso de expansión...")

    progress_bar.progress(10, text="Paso 1 / 5 — Cargando configuración y mapeo de menciones...")
    region_map, internet_map, mention_map = load_config_maps(config_file, include_mentions=True)

    progress_bar.progress(25, text="Paso 2 / 5 — Leyendo Dossier...")
    df = read_dossier(dossier_file)

    progress_bar.progress(40, text="Paso 3 / 5 — Aplicando mapeos y normalizaciones...")
    df = apply_common_transformations(df, region_map, internet_map)

    progress_bar.progress(55, text="Paso 4 / 5 — Mapeando menciones...")
    df = apply_mention_mapping(df, mention_map)

    progress_bar.progress(75, text="Paso 5 / 5 — Expandiendo filas por menciones...")
    original_count = len(df)
    df_expanded = utils.expand_by_mentions(df, 'Menciones - Empresa')
    expanded_count = len(df_expanded)
    new_rows = expanded_count - original_count

    progress_bar.progress(100, text="✓ Expansión completada")

    filename   = f"Dossier_Expandido_{datetime.datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    excel_data = utils.to_excel_from_df(df_expanded, FINAL_ORDER)

    with download_placeholder:
        st.download_button(
            label="⬇ Descargar archivo expandido (.xlsx)",
            data=excel_data,
            file_name=filename,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key="download_expanded"
        )

    unique_mentions = df_expanded['Menciones - Empresa'].nunique()

    st.markdown(f"""
    <div class="success-banner">
        <div class="icon">📋</div>
        <div class="text">
            <strong>Expansión finalizada correctamente</strong>
            <span>{original_count:,} filas originales → {expanded_count:,} filas expandidas (+{new_rows:,} nuevas)</span>
        </div>
    </div>
    """, unsafe_allow_html=True)

    st.markdown('<p class="results-header">Resumen de la expansión</p>', unsafe_allow_html=True)
    st.markdown(f"""
    <div class="metrics-row">
        <div class="metric-card">
            <div class="metric-value">{original_count:,}</div>
            <div class="metric-label">Filas originales</div>
        </div>
        <div class="metric-card accent">
            <div class="metric-value">{expanded_count:,}</div>
            <div class="metric-label">Filas después de expandir</div>
        </div>
        <div class="metric-card muted">
            <div class="metric-value">{unique_mentions:,}</div>
            <div class="metric-label">Menciones únicas detectadas</div>
        </div>
    </div>
    """, unsafe_allow_html=True)

    with st.expander("📊  Ver desglose por mención"):
        mention_counts = df_expanded['Menciones - Empresa'].value_counts().reset_index()
        mention_counts.columns = ['Mención', 'Cantidad de filas']
        st.dataframe(mention_counts, use_container_width=True, hide_index=True)

    st.markdown('<p class="results-header">Previsualización de resultados</p>', unsafe_allow_html=True)
    final_cols_in_df = [col for col in FINAL_ORDER if col in df_expanded.columns]
    df_display = df_expanded[final_cols_in_df].copy()
    if 'Fecha' in df_display.columns:
        df_display['Fecha'] = (
            pd.to_datetime(df_display['Fecha'])
            .dt.strftime('%d/%m/%Y')
            .replace('NaT', 'FECHA INVÁLIDA')
        )
    st.dataframe(df_display, use_container_width=True, hide_index=True)


# ==============================================================================
# INTERFAZ PRINCIPAL
# ==============================================================================

st.markdown("""
<div class="app-header">
    <div class="badge">Transmilenio · Media Intelligence</div>
    <p>Limpieza, homogeneización IA y análisis automático de dossiers · v1.6 | Johnathan Cortés 😼</p>
</div>
""", unsafe_allow_html=True)

tab_process, tab_expand = st.tabs(["🔄 Procesar Dossier", "📋 Expandir por Menciones"])


# ── PESTAÑA 1: PROCESO COMPLETO ──
with tab_process:
    st.markdown("""
    <div class="card">
        <div class="card-title">Cómo usar esta herramienta</div>
        <div class="step">
            <div class="step-num">1</div>
            <div class="step-text">Prepara tu archivo <strong>Dossier</strong> (.xlsx) y el archivo <strong>Configuracion.xlsx</strong>.</div>
        </div>
        <div class="step">
            <div class="step-num">2</div>
            <div class="step-text">Sube ambos archivos en el área de carga. El sistema los detecta automáticamente.</div>
        </div>
        <div class="step">
            <div class="step-num">3</div>
            <div class="step-text">Haz clic en <strong>Iniciar proceso</strong>. Al finalizar, el botón de descarga aparecerá a su lado.</div>
        </div>
    </div>
    """, unsafe_allow_html=True)

    with st.expander("📋  Ver estructura requerida para Configuracion.xlsx"):
        st.markdown("""
        | Hoja | Columna A | Columna B |
        |------|-----------|-----------|
        | `Regiones` | Medio | Región |
        | `Internet` | Medio Original | Medio Mapeado |

        > **Nota:** En esta pestaña la columna `Menciones - Empresa` se mantiene tal cual viene en el dossier.
        """)

    st.markdown("<br>", unsafe_allow_html=True)
    st.markdown('<div class="card-title">Carga de archivos</div>', unsafe_allow_html=True)

    uploaded_files_p = st.file_uploader(
        "Arrastra los archivos aquí",
        type=["xlsx"],
        accept_multiple_files=True,
        label_visibility="collapsed",
        key="uploader_process"
    )

    dossier_file_p, config_file_p = None, None
    if uploaded_files_p:
        for file in uploaded_files_p:
            if 'config' in file.name.lower():
                config_file_p = file
            else:
                dossier_file_p = file

        col_a, col_b = st.columns(2)
        with col_a:
            if dossier_file_p:
                st.markdown(f'<div class="file-status ok">✓ Dossier — <strong>{dossier_file_p.name}</strong></div>', unsafe_allow_html=True)
            else:
                st.markdown('<div class="file-status missing">⚠ No se detectó el archivo Dossier</div>', unsafe_allow_html=True)
        with col_b:
            if config_file_p:
                st.markdown(f'<div class="file-status ok">✓ Configuración — <strong>{config_file_p.name}</strong></div>', unsafe_allow_html=True)
            else:
                st.markdown('<div class="file-status missing">⚠ No se detectó Configuracion.xlsx</div>', unsafe_allow_html=True)

    st.markdown("<br>", unsafe_allow_html=True)

    col_start_p, col_dl_p = st.columns(2)
    with col_start_p:
        start_p = st.button(
            "▶  Iniciar proceso completo",
            disabled=not (dossier_file_p and config_file_p),
            type="primary",
            key="btn_start_process"
        )
    with col_dl_p:
        dl_placeholder_p = st.empty()
        if not start_p:
            with dl_placeholder_p:
                st.button("⬇ Descargar archivo procesado (.xlsx)", disabled=True, type="primary", key="btn_dl_p_ph")

    if start_p:
        run_full_process(dossier_file_p, config_file_p, dl_placeholder_p)


# ── PESTAÑA 2: EXPANDIR POR MENCIONES ──
with tab_expand:
    st.markdown("""
    <div class="card">
        <div class="card-title">Expandir filas por menciones</div>
        <div class="step">
            <div class="step-num">1</div>
            <div class="step-text">Prepara tu archivo <strong>Dossier</strong> (.xlsx) y el archivo <strong>Configuracion.xlsx</strong> (debe incluir la hoja <code>Menciones</code>).</div>
        </div>
        <div class="step">
            <div class="step-num">2</div>
            <div class="step-text">Sube ambos archivos. Se aplican mapeos de regiones, medios internet <strong>y menciones</strong>.</div>
        </div>
        <div class="step">
            <div class="step-num">3</div>
            <div class="step-text">Las filas con múltiples menciones separadas por <strong>;</strong> en <code>Menciones - Empresa</code> se duplican/triplican. <strong>No se predicen tono ni tema.</strong></div>
        </div>
    </div>
    """, unsafe_allow_html=True)

    with st.expander("📋  Ver estructura requerida para Configuracion.xlsx"):
        st.markdown("""
        | Hoja | Columna A | Columna B |
        |------|-----------|-----------|
        | `Regiones` | Medio | Región |
        | `Internet` | Medio Original | Medio Mapeado |
        | `Menciones` | Mención Original | Mención Mapeada |

        > La hoja `Menciones` es **obligatoria** en esta pestaña para normalizar los nombres antes de expandir.
        """)

    st.markdown("<br>", unsafe_allow_html=True)
    st.markdown('<div class="card-title">Carga de archivos</div>', unsafe_allow_html=True)

    uploaded_files_e = st.file_uploader(
        "Arrastra los archivos aquí",
        type=["xlsx"],
        accept_multiple_files=True,
        label_visibility="collapsed",
        key="uploader_expand"
    )

    dossier_file_e, config_file_e = None, None
    if uploaded_files_e:
        for file in uploaded_files_e:
            if 'config' in file.name.lower():
                config_file_e = file
            else:
                dossier_file_e = file

        col_a2, col_b2 = st.columns(2)
        with col_a2:
            if dossier_file_e:
                st.markdown(f'<div class="file-status ok">✓ Dossier — <strong>{dossier_file_e.name}</strong></div>', unsafe_allow_html=True)
            else:
                st.markdown('<div class="file-status missing">⚠ No se detectó el archivo Dossier</div>', unsafe_allow_html=True)
        with col_b2:
            if config_file_e:
                st.markdown(f'<div class="file-status ok">✓ Configuración — <strong>{config_file_e.name}</strong></div>', unsafe_allow_html=True)
            else:
                st.markdown('<div class="file-status missing">⚠ No se detectó Configuracion.xlsx</div>', unsafe_allow_html=True)

    st.markdown("<br>", unsafe_allow_html=True)

    col_start_e, col_dl_e = st.columns(2)
    with col_start_e:
        start_e = st.button(
            "▶  Expandir por menciones",
            disabled=not (dossier_file_e and config_file_e),
            type="primary",
            key="btn_start_expand"
        )
    with col_dl_e:
        dl_placeholder_e = st.empty()
        if not start_e:
            with dl_placeholder_e:
                st.button("⬇ Descargar archivo expandido (.xlsx)", disabled=True, type="primary", key="btn_dl_e_ph")

    if start_e:
        run_expand_process(dossier_file_e, config_file_e, dl_placeholder_e)
